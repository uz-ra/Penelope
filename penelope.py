#!/usr/bin/env python3
import argparse
import csv
import json
import re
import subprocess
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FIELD_MAP = {
    "name": "表記する名前",
    "title": "作品タイトル / Title",
    "university": "所属大学 / University",
    "major": "学部・学科 / Major",
    "grade": "学年 / Grade",
    "camera": "使用機材 / Camera",
    "lens": "使用レンズ / Lens",
    "shutter": "シャッタスピード / Shutter speed",
    "aperture": "絞り値 / Aperture",
    "iso": "ISO感度（使用フィルム） / ISO(Film name)",
    "location": "撮影場所 / Location",
    "caption": "キャプション / Caption",
}

CELL_MAP_WITH_CAPTION = {
    "title": "C3",
    "location": "R3",
    "affiliation": "C6",
    "name": "C7",
    "camera": "R6",
    "lens": "R9",
    "aperture": "R12",
    "shutter": "V12",
    "iso": "Z12",
    "caption": "C10",
}

CELL_MAP_NO_CAPTION = {
    "title": "C5",
    "location": "R3",
    "affiliation": "C9",
    "name": "C10",
    "camera": "R6",
    "lens": "R9",
    "aperture": "R12",
    "shutter": "V12",
    "iso": "Z12",
}

TEMPLATE_MIN_COL = 1
TEMPLATE_MAX_COL = 30
TEMPLATE_MIN_ROW = 1
TEMPLATE_MAX_ROW = 14
CAPTION_WIDTH_PX = 480
CAPTION_HEIGHT_PX = 224
CAPTION_TRIM_TOLERANCE = 1
A4_WIDTH_MM = 210
CAPTION_EMPTY_PHRASES = ("キャプション無し", "キャプションなし")
PRINT_AREA = f"A1:{get_column_letter(TEMPLATE_MAX_COL)}{TEMPLATE_MAX_ROW}"
PACKING_GAP = 0.0
DEFAULT_COL_WIDTH = 8.43
STATIC_TEXT_CELLS = {"U12", "Y12"}


def clear_block_values(
    ws,
    min_col: int,
    max_col: int,
    min_row: int,
    max_row: int,
    keep_cells: set[str] | None = None,
) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue
            if keep_cells and cell.coordinate in keep_cells:
                continue
            if cell.coordinate in ws.merged_cells:
                is_start = False
                for merge_range in ws.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        is_start = cell.coordinate == merge_range.start_cell.coordinate
                        break
                if not is_start:
                    continue
            cell.value = None



def _mm_to_points(mm: float) -> float:
    return mm * 72 / 25.4


def _default_row_height(ws) -> float:
    return ws.sheet_format.defaultRowHeight or 15


def _row_height(ws, row_index: int) -> float:
    height = ws.row_dimensions[row_index].height
    if height is None:
        return _default_row_height(ws)
    return height


def _column_width_points(ws, col_index: int) -> float:
    col_letter = get_column_letter(col_index)
    width = ws.column_dimensions[col_letter].width
    if width is None:
        width = ws.sheet_format.defaultColWidth or DEFAULT_COL_WIDTH
    if width <= 1:
        pixels = width * 12 + 0.5
    else:
        pixels = width * 7 + 5
    return pixels * 72 / 96


def _sheet_size_points(ws, max_col: int, max_row: int) -> tuple[float, float]:
    total_width = 0.0
    for col_index in range(TEMPLATE_MIN_COL, max_col + 1):
        total_width += _column_width_points(ws, col_index)

    total_height = 0.0
    for row_index in range(TEMPLATE_MIN_ROW, max_row + 1):
        height = ws.row_dimensions[row_index].height
        if height is None:
            height = _default_row_height(ws)
        total_height += height

    return total_width, total_height


def _print_bounds(ws) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row in ws.iter_rows(min_row=TEMPLATE_MIN_ROW, max_row=TEMPLATE_MAX_ROW,
                            min_col=TEMPLATE_MIN_COL, max_col=TEMPLATE_MAX_COL):
        row_has_border = False
        row_index = row[0].row
        for cell in row:
            border = cell.border
            if any(side.style for side in (border.left, border.right, border.top, border.bottom)):
                row_has_border = True
                max_col = max(max_col, cell.column)
            if cell.value is not None:
                row_has_border = True
                max_col = max(max_col, cell.column)
        if row_has_border:
            max_row = max(max_row, row_index)

    if max_row == 0:
        max_row = TEMPLATE_MAX_ROW
    if max_col == 0:
        max_col = TEMPLATE_MAX_COL

    return max_col, max_row


def normalize_caption(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    for phrase in CAPTION_EMPTY_PHRASES:
        if phrase in value:
            return ""
    return value


def normalize_optional_text(value: str) -> str:
    value = (value or "").strip()
    if not value or value in {"-", "ー", "ｰ"}:
        return ""
    return value


def normalize_affiliation_part(value: str) -> str:
    value = normalize_optional_text(value)
    if not value:
        return ""
    return value.split("/", 1)[0].strip()


def normalize_shutter(value: str) -> str:
    value = normalize_optional_text(value)
    if not value:
        return ""
    if re.fullmatch(r"[0-9./]+", value):
        return f"{value}s"
    return value


def normalize_aperture(value: str) -> str:
    value = normalize_optional_text(value)
    if not value:
        return ""
    if value[0] in {"F", "f", "ƒ"}:
        return f"ƒ{value[1:]}" if value[0] != "ƒ" else value
    return f"ƒ{value}"


def normalize_iso(value: str) -> str:
    value = normalize_optional_text(value)
    if not value:
        return ""
    upper = value.upper()
    if upper.startswith("ISO"):
        return value
    if any(ch.isdigit() for ch in value) and not any(ch.isalpha() for ch in value):
        return f"ISO {value}"
    return value


def load_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        missing = [name for name in FIELD_MAP.values() if name not in reader.fieldnames]
        if missing:
            missing_text = ", ".join(missing)
            raise ValueError(f"CSVに必要な列がありません: {missing_text}")
        rows = []
        for row in reader:
            rows.append(row)
        return rows


def extract_values(row: dict[str, str]) -> dict[str, str]:
    affiliation_parts = [
        normalize_affiliation_part(row.get(FIELD_MAP["university"])),
        normalize_affiliation_part(row.get(FIELD_MAP["major"])),
        normalize_affiliation_part(row.get(FIELD_MAP["grade"])),
    ]
    affiliation = " ".join(part for part in affiliation_parts if part)
    return {
        "name": (row.get(FIELD_MAP["name"]) or "").strip(),
        "title": (row.get(FIELD_MAP["title"]) or "").strip(),
        "affiliation": affiliation,
        "camera": (row.get(FIELD_MAP["camera"]) or "").strip(),
        "lens": (row.get(FIELD_MAP["lens"]) or "").strip(),
        "aperture": normalize_aperture(row.get(FIELD_MAP["aperture"])),
        "shutter": normalize_shutter(row.get(FIELD_MAP["shutter"])),
        "iso": normalize_iso(row.get(FIELD_MAP["iso"])),
        "location": (row.get(FIELD_MAP["location"]) or "").strip(),
        "caption": normalize_caption(row.get(FIELD_MAP["caption"])),
    }


def fill_caption(ws, values: dict[str, str], cell_map: dict[str, str]) -> None:
    if not any(values.values()):
        return

    for key, cell in cell_map.items():
        ws[cell].value = values.get(key) if values.get(key) else None


def apply_print_settings(ws) -> None:
    ws.print_area = PRINT_AREA
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    a4_width_points = _mm_to_points(A4_WIDTH_MM)
    height_ratio = CAPTION_HEIGHT_PX / CAPTION_WIDTH_PX
    paper_height_points = a4_width_points * height_ratio
    ws.page_setup.paperSize = 0
    ws.page_setup.paperWidth = f"{a4_width_points / 72:.3f}in"
    ws.page_setup.paperHeight = f"{paper_height_points / 72:.3f}in"
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def export_pdf(xlsx_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{xlsx_path.stem}.pdf"

    def escape_applescript(value: str) -> str:
        return value.replace("\\", "\\\\").replace('"', '\\"')

    xlsx_str = escape_applescript(str(xlsx_path))
    pdf_str = escape_applescript(str(output_path))

    script_lines = [
        'tell application "Microsoft Excel"',
        'activate',
        f'set wb to open workbook workbook file name (POSIX file "{xlsx_str}")',
        f'set pdf_file to POSIX file "{pdf_str}"',
        'save workbook as wb filename pdf_file file format PDF file format',
        'close wb saving no',
        'end tell',
    ]

    try:
        command = ["osascript"]
        for line in script_lines:
            command.extend(["-e", line])
        subprocess.run(command, check=True)
    except FileNotFoundError as exc:
        raise SystemExit("osascriptが見つかりません。macOSで実行してください。") from exc
    except subprocess.CalledProcessError as exc:
        raise SystemExit("ExcelでのPDF生成に失敗しました。Excelが起動できるか確認してください。未起動の場合はExcelを起動してください。") from exc

    if not output_path.exists():
        raise SystemExit("PDFの生成に失敗しました。")

    return output_path


def _page_has_content(page) -> bool:
    contents = page.get_contents()
    if contents is None:
        return False
    try:
        data = contents.get_data()
    except Exception:
        return True
    return bool(data and data.strip())


def _multiply_matrices(m1: list[float], m2: list[float]) -> list[float]:
    a1, b1, c1, d1, e1, f1 = m1
    a2, b2, c2, d2, e2, f2 = m2
    return [
        a1 * a2 + c1 * b2,
        b1 * a2 + d1 * b2,
        a1 * c2 + c1 * d2,
        b1 * c2 + d1 * d2,
        a1 * e2 + c1 * f2 + e1,
        b1 * e2 + d1 * f2 + f1,
    ]


def _apply_matrix(m: list[float], x: float, y: float) -> tuple[float, float]:
    a, b, c, d, e, f = m
    return (a * x + c * y + e, b * x + d * y + f)


def _find_border_rect(page):
    try:
        from pypdf.generic import ContentStream
    except ImportError:
        return None

    contents = page.get_contents()
    if contents is None:
        return None

    stream = ContentStream(contents, page.pdf)
    ctm_stack = [[1, 0, 0, 1, 0, 0]]
    rects: list[tuple[float, float, float, float]] = []

    for operands, operator in stream.operations:
        if operator == b"q":
            ctm_stack.append(ctm_stack[-1][:])
        elif operator == b"Q":
            if len(ctm_stack) > 1:
                ctm_stack.pop()
        elif operator == b"cm":
            matrix = [float(v) for v in operands]
            ctm_stack[-1] = _multiply_matrices(ctm_stack[-1], matrix)
        elif operator == b"re":
            x, y, w, h = [float(v) for v in operands]
            points = [
                _apply_matrix(ctm_stack[-1], x, y),
                _apply_matrix(ctm_stack[-1], x + w, y),
                _apply_matrix(ctm_stack[-1], x, y + h),
                _apply_matrix(ctm_stack[-1], x + w, y + h),
            ]
            xs = [pt[0] for pt in points]
            ys = [pt[1] for pt in points]
            rects.append((min(xs), min(ys), max(xs), max(ys)))

    if not rects:
        return None

    rects.sort(key=lambda r: (r[2] - r[0]) * (r[3] - r[1]), reverse=True)
    largest = rects[0]
    largest_area = (largest[2] - largest[0]) * (largest[3] - largest[1])
    largest_cx = (largest[0] + largest[2]) / 2
    largest_cy = (largest[1] + largest[3]) / 2

    candidates = []
    for rect in rects[1:]:
        area = (rect[2] - rect[0]) * (rect[3] - rect[1])
        if area < largest_area * 0.9 or area >= largest_area:
            continue
        cx = (rect[0] + rect[2]) / 2
        cy = (rect[1] + rect[3]) / 2
        if abs(cx - largest_cx) > 5 or abs(cy - largest_cy) > 5:
            continue
        candidates.append(rect)

    if candidates:
        candidates.sort(key=lambda r: (r[2] - r[0]) * (r[3] - r[1]))
        inner = candidates[0]
        mid = (
            (largest[0] + inner[0]) / 2,
            (largest[1] + inner[1]) / 2,
            (largest[2] + inner[2]) / 2,
            (largest[3] + inner[3]) / 2,
        )
        return mid

    return largest


def _trim_rect_to_caption_size(rect: tuple[float, float, float, float]) -> tuple[float, float, float, float]:
    x0, y0, x1, y1 = rect
    width = x1 - x0
    height = y1 - y0
    if width < CAPTION_WIDTH_PX - CAPTION_TRIM_TOLERANCE:
        return rect
    if height < CAPTION_HEIGHT_PX - CAPTION_TRIM_TOLERANCE:
        return rect
    if width - CAPTION_WIDTH_PX > CAPTION_TRIM_TOLERANCE or height - CAPTION_HEIGHT_PX > CAPTION_TRIM_TOLERANCE:
        x1 = x0 + CAPTION_WIDTH_PX
        y0 = y1 - CAPTION_HEIGHT_PX
    return (x0, y0, x1, y1)


def _trim_page_to_bbox(page):
    try:
        from pypdf import PageObject, Transformation
        from pypdf.generic import RectangleObject
    except ImportError:
        return None

    rect = _find_border_rect(page)
    if rect is None:
        box = page.cropbox
        rect = (float(box.lower_left[0]), float(box.lower_left[1]), float(box.upper_right[0]), float(box.upper_right[1]))
    rect = _trim_rect_to_caption_size(rect)

    rect_obj = RectangleObject(rect)
    width = float(rect_obj.width)
    height = float(rect_obj.height)

    trimmed = PageObject.create_blank_page(width=width, height=height)
    page_copy = copy(page)
    page_copy.cropbox.lower_left = rect_obj.lower_left
    page_copy.cropbox.upper_right = rect_obj.upper_right
    page_copy.mediabox.lower_left = rect_obj.lower_left
    page_copy.mediabox.upper_right = rect_obj.upper_right
    trimmed.merge_transformed_page(
        page_copy,
        Transformation().translate(tx=-rect_obj.lower_left[0], ty=-rect_obj.lower_left[1]),
    )
    return trimmed


def split_pdf_pages(source_pdf: Path, output_dir: Path, prefix: str = "caption_") -> list:
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError as exc:
        raise SystemExit("pypdfが必要です。`python -m pip install pypdf` を実行してください。") from exc

    reader = PdfReader(str(source_pdf))
    output_dir.mkdir(parents=True, exist_ok=True)
    pages = []

    index = 1
    for page in reader.pages:
        if not _page_has_content(page):
            continue
        trimmed = _trim_page_to_bbox(page)
        if trimmed is None:
            continue
        writer = PdfWriter()
        writer.add_page(trimmed)
        path = output_dir / f"{prefix}{index:03d}.pdf"
        with path.open("wb") as f:
            writer.write(f)
        pages.append(trimmed)
        index += 1

    return pages


def make_2up_pdf(source_pdf: Path, output_pdf: Path, pages_override: list | None = None) -> None:
    try:
        from pypdf import PdfReader, PdfWriter, PageObject, Transformation
    except ImportError as exc:
        raise SystemExit("pypdfが必要です。`python -m pip install pypdf` を実行してください。") from exc

    if pages_override is None:
        reader = PdfReader(str(source_pdf))
        pages = [page for page in reader.pages if _page_has_content(page)]
    else:
        pages = pages_override
    if not pages:
        raise SystemExit("PDFに有効なページがありません。")

    writer = PdfWriter()
    base_width = _mm_to_points(210)
    base_height = _mm_to_points(297)

    current_page = PageObject.create_blank_page(width=base_width, height=base_height)
    y_position = base_height
    has_content = False

    for page in pages:
        trimmed = _trim_page_to_bbox(page) if pages_override is None else page
        if trimmed is None:
            continue

        box = trimmed.mediabox
        width = float(box.width)
        height = float(box.height)
        scale = base_width / width
        scaled_height = height * scale
        if scaled_height > base_height and height:
            scale = base_height / height
            scaled_height = height * scale
        required_height = scaled_height + PACKING_GAP

        if required_height > y_position:
            writer.add_page(current_page)
            current_page = PageObject.create_blank_page(width=base_width, height=base_height)
            y_position = base_height
            has_content = False

        y_position -= required_height
        x_offset = 0
        y_offset = y_position

        page_copy = copy(trimmed)
        current_page.merge_transformed_page(
            page_copy,
            Transformation().scale(sx=scale, sy=scale).translate(tx=x_offset, ty=y_offset),
        )
        has_content = True

        if y_position <= 0:
            writer.add_page(current_page)
            current_page = PageObject.create_blank_page(width=base_width, height=base_height)
            y_position = base_height
            has_content = False

    if current_page is not None and has_content:
        writer.add_page(current_page)

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    with output_pdf.open("wb") as f:
        writer.write(f)


def resolve_path(base_dir: Path, value: str | None) -> Path | None:
    if not value:
        return None
    path = Path(value)
    if not path.is_absolute():
        path = base_dir / path
    return path


def load_config(config_path: Path) -> dict[str, str | None]:
    if not config_path.exists():
        raise SystemExit(f"config.jsonが見つかりません: {config_path}")

    with config_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    csv_path = data.get("csv")
    template_path = data.get("template")
    if not csv_path or not template_path:
        raise SystemExit("config.jsonにcsvとtemplateを指定してください")

    output_dir = data.get("output_dir") or "examples"
    output_xlsx = data.get("output_xlsx") or "captions.xlsx"
    output_pdf = data.get("output_pdf")
    output_pdf_2up = data.get("output_pdf_2up")
    output_pdf_dir = data.get("output_pdf_dir")

    return {
        "csv": csv_path,
        "template": template_path,
        "output_dir": output_dir,
        "output_xlsx": output_xlsx,
        "output_pdf": output_pdf,
        "output_pdf_2up": output_pdf_2up,
        "output_pdf_dir": output_pdf_dir,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="CSVからキャプション印刷用XLSXを作成します")
    parser.add_argument("--config", help="設定ファイルのパス（デフォルト: config.json）")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent
    config_path = resolve_path(base_dir, args.config) if args.config else base_dir / "config.json"
    config = load_config(config_path)

    csv_path = resolve_path(base_dir, config["csv"])
    template_path = resolve_path(base_dir, config["template"])
    output_dir = resolve_path(base_dir, config["output_dir"])
    output_xlsx = Path(config["output_xlsx"]) if config["output_xlsx"] else None
    output_pdf = Path(config["output_pdf"]) if config["output_pdf"] else None
    output_pdf_2up = Path(config["output_pdf_2up"]) if config["output_pdf_2up"] else None
    output_pdf_dir = Path(config["output_pdf_dir"]) if config.get("output_pdf_dir") else None

    if csv_path is None or template_path is None or output_dir is None or output_xlsx is None:
        raise SystemExit("config.jsonのパス指定が不正です")

    rows = load_rows(csv_path)
    if not rows:
        raise SystemExit("CSVにデータ行がありません")

    wb = load_workbook(template_path)
    template_with_caption = wb.sheetnames[0]
    template_no_caption = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
    ws_with_caption = wb[template_with_caption]
    ws_no_caption = wb[template_no_caption]

    for index, row in enumerate(rows, start=1):
        values = extract_values(row)
        use_no_caption = not values.get("caption")
        base_template = ws_no_caption if use_no_caption else ws_with_caption
        cell_map = CELL_MAP_NO_CAPTION if use_no_caption else CELL_MAP_WITH_CAPTION

        ws = wb.copy_worksheet(base_template)
        ws.title = f"Caption_{index:03d}"
        clear_block_values(
            ws,
            TEMPLATE_MIN_COL,
            TEMPLATE_MAX_COL,
            TEMPLATE_MIN_ROW,
            TEMPLATE_MAX_ROW,
            keep_cells=STATIC_TEXT_CELLS,
        )
        fill_caption(ws, values, cell_map)
        apply_print_settings(ws)

    wb.remove(ws_with_caption)
    if ws_no_caption is not ws_with_caption:
        wb.remove(ws_no_caption)

    output_dir.mkdir(parents=True, exist_ok=True)
    if not output_xlsx.is_absolute():
        output_xlsx = output_dir / output_xlsx
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    if output_pdf and not output_pdf.is_absolute():
        output_pdf = output_dir / output_pdf
    if output_pdf_2up and not output_pdf_2up.is_absolute():
        output_pdf_2up = output_dir / output_pdf_2up
    if output_pdf_dir and not output_pdf_dir.is_absolute():
        output_pdf_dir = output_dir / output_pdf_dir

    if output_pdf or output_pdf_2up or output_pdf_dir:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_pdf = export_pdf(output_xlsx, Path(tmp_dir))
            source_pdf = temp_pdf

            trimmed_pages = None
            if output_pdf_dir:
                trimmed_pages = split_pdf_pages(source_pdf, output_pdf_dir)

            if output_pdf:
                output_pdf.parent.mkdir(parents=True, exist_ok=True)
                temp_pdf.replace(output_pdf)
                source_pdf = output_pdf

            if output_pdf_2up:
                make_2up_pdf(source_pdf, output_pdf_2up, pages_override=trimmed_pages)


if __name__ == "__main__":
    main()
